#!/usr/bin/env python3
"""
TRES Finance Report Analyzer
Reads any TRES report XLSX and outputs structured findings as JSON.

Usage:
    python analyze_report.py <xlsx_path> --output <output.json>
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(
        "Error: openpyxl is not installed.\n"
        "Please run: python3 -m venv .venv && .venv/bin/pip install openpyxl==3.1.5",
        file=sys.stderr,
    )
    sys.exit(1)


def read_sheet_as_dicts(ws):
    """Read a worksheet into a list of dicts using the first row as headers.
    Handles pivot tables where the real headers may be in row 2 (row 1 is just 'Values')."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    # Check if first row looks like a pivot header (single value like "Values")
    first_row = rows[0]
    non_none_first = [v for v in first_row if v is not None]
    if len(non_none_first) <= 1 and len(rows) > 2:
        # Likely a pivot table -- check if row 2 has more column names
        second_row = rows[1]
        non_none_second = [v for v in second_row if v is not None]
        if len(non_none_second) > len(non_none_first):
            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(second_row)]
            data = []
            for row in rows[2:]:
                if all(v is None for v in row):
                    continue
                data.append(dict(zip(headers, row)))
            return headers, data

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(first_row)]
    data = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        data.append(dict(zip(headers, row)))
    return headers, data


def safe_float(val, default=0.0):
    """Convert a value to float safely."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def identify_report_type(wb):
    """Identify the TRES report type from sheet names and columns."""
    sheet_names = wb.sheetnames
    sheet_set = set(s.strip() for s in sheet_names)

    # Ledger Reconciliation: 7 specific tabs
    recon_tabs = {"Summary", "Inventory Reconciliation", "Running Token Reconciliation",
                  "Running Fiat Reconciliation", "Historical Token Reconciliation",
                  "Historical Fiat Reconciliation", "Roll Forward Reconciliation"}
    if recon_tabs.issubset(sheet_set):
        return "RECONCILIATION_LEDGER"

    # Realized Gains & Losses: 4 specific tabs
    if {"Summary Per Asset", "Summary Per Year", "Summary per Tx Activity"}.issubset(sheet_set):
        return "EXTENDED_RAW_TRANSACTIONS"

    # Balance Trends vs Asset Balances: both have By Asset, By Wallet, etc.
    pivot_tabs = {"By Asset", "By Wallet", "By Platform", "By Position", "Cost Basis"}
    if pivot_tabs.issubset(sheet_set):
        if "raw_data" in sheet_set:
            headers = [str(c.value).strip() for c in list(wb["raw_data"].iter_rows(max_row=1))[0] if c.value]
            if any("Previous" in h for h in headers):
                return "BALANCE_TRENDS"
            # Historical Balance Format: same tabs as Asset Balances but with Historical Balance columns
            if any("Historical Balance" in h for h in headers):
                return "HISTORICAL_BALANCE"
        return "RAW_BALANCES"

    # Asset Balances V2
    if "Asset Balances - PT" in sheet_set:
        return "RAW_BALANCES_V2"

    # Archives
    if {"Fiat Value Summary", "Amount Summary By Application", "Amount Summary"}.issubset(sheet_set):
        return "ARCHIVED_BALANCES"

    # Asset Roll Forward
    if "Overview" in sheet_set and "raw_data" in sheet_set:
        headers = [str(c.value).strip() for c in list(wb["raw_data"].iter_rows(max_row=1))[0] if c.value]
        if any("Safety Check" in h for h in headers):
            return "ASSET_ROLL_FORWARD"

    # Cost Basis Roll Forward
    if "Summary" in sheet_set and "Inventory Reconciliation" in sheet_set and "raw_data" in sheet_set:
        return "COST_BASIS_ROLL_FORWARD"

    # ERP Pre-Sync
    if "Chart of Accounts Summary" in sheet_set:
        return "PRE_SYNC_JOURNAL"

    # Single raw_data tab -- need to check columns
    if "raw_data" in sheet_set:
        headers = [str(c.value).strip() for c in list(wb["raw_data"].iter_rows(max_row=1))[0] if c.value]
        header_set = set(headers)

        if any("COGS Lot" in h for h in headers):
            return "COST_BREAKDOWN_RAW_TRANSACTIONS"
        if any("Rollup Parent" in h for h in headers):
            return "ROLLUP_BREAKDOWN"
        if "Sync Status" in header_set and len(headers) < 12:
            return "POST_SYNC_JOURNAL"
        if "Price Source" in header_set:
            return "DAILY_ASSET_PRICING"
        if "Purchase Date" in header_set and "Remaining Quantity" in header_set:
            return "COST_BASIS_STACK_PER_ACCOUNT"
        if "Sub TX Index" in header_set and "Is Taxable" in header_set:
            return "COST_BASIS_INVENTORY"
        if "Status" in header_set and "Added Date" in header_set:
            return "INTERNAL_ACCOUNTS"
        # Default: Transaction Ledger
        if any("TX Hash" in h or "Tx Hash" in h for h in headers):
            return "BASIC_RAW_TRANSACTIONS"

    return "UNKNOWN"


def analyze_reconciliation_ledger(wb):
    """Analyze Ledger Reconciliation report."""
    findings = {"report_type": "RECONCILIATION_LEDGER", "report_name": "Ledger Reconciliation"}

    # Summary tab -- note: pivot tabs often have formula-based values that
    # appear as None when the XLSX was not opened in Excel before export.
    # We read what we can but rely on raw_data for the real analysis.
    ws = wb["Summary"]
    summary_data = {}
    has_values = False
    for row in ws.iter_rows(values_only=True):
        if row and row[0]:
            key = str(row[0]).strip()
            vals = {
                "running_fiat": safe_float(row[1]) if len(row) > 1 else 0,
                "running_token": safe_float(row[2]) if len(row) > 2 else 0,
                "historical_fiat": safe_float(row[3]) if len(row) > 3 else 0,
                "historical_token": safe_float(row[4]) if len(row) > 4 else 0,
            }
            if any(v != 0 for v in vals.values()):
                has_values = True
            summary_data[key] = vals
    if has_values:
        findings["summary"] = summary_data
    else:
        findings["summary_note"] = "Summary tab contains formulas without cached values. Analysis is based on raw_data instead."

    # Roll Forward Reconciliation - most granular
    headers, data = read_sheet_as_dicts(wb["Roll Forward Reconciliation"])
    if data:
        diff_col = None
        for h in headers:
            if "Historical Close" in h and "Expected" in h:
                diff_col = h
                break
        if not diff_col:
            for h in headers:
                if "Expected Close" in h and h != next((x for x in headers if x.endswith("Expected Close (T)")), ""):
                    diff_col = h
                    break
        # Also try the exact TRES column name with "Sum of" prefix
        if not diff_col:
            for h in headers:
                if "Close - Expected" in h or "Expected Close" in h:
                    if "Sum of" in h or "Historical" in h:
                        diff_col = h
                        break

        if diff_col:
            failing = [r for r in data if safe_float(r.get(diff_col)) != 0]
            findings["roll_forward"] = {
                "total_balance_ids": len(data),
                "passing": len(data) - len(failing),
                "failing": len(failing),
                "pass_rate": round((len(data) - len(failing)) / len(data) * 100, 1) if data else 0,
                "top_failures": sorted(
                    [{"wallet": r.get("Wallet Name", "?"), "asset": r.get("Asset Name", "?"),
                      "gap": safe_float(r.get(diff_col))} for r in failing],
                    key=lambda x: abs(x["gap"]), reverse=True
                )[:10]
            }

    # Historical Token Reconciliation
    headers, data = read_sheet_as_dicts(wb["Historical Token Reconciliation"])
    if data:
        check_open_col = next((h for h in headers if "Check Open" in h), None)
        check_close_col = next((h for h in headers if "Check Close" in h), None)
        open_drift = [r for r in data if check_open_col and safe_float(r.get(check_open_col)) != 0]
        close_drift = [r for r in data if check_close_col and safe_float(r.get(check_close_col)) != 0]
        findings["historical_token"] = {
            "total_assets": len(data),
            "open_drift_count": len(open_drift),
            "close_drift_count": len(close_drift),
        }

    # Check raw_data for slippage and missing prices
    if "raw_data" in wb.sheetnames:
        headers, data = read_sheet_as_dicts(wb["raw_data"])
        if data:
            slip_t = [r for r in data if safe_float(r.get("Slippage (T)", 0)) != 0]
            slip_f = [r for r in data if safe_float(r.get("Slippage ($)", 0)) != 0]
            no_price = [r for r in data if str(r.get("Has Price?", "")).lower() in ("false", "no", "0")]
            findings["raw_data"] = {
                "total_rows": len(data),
                "slippage_token_count": len(slip_t),
                "slippage_fiat_count": len(slip_f),
                "missing_price_count": len(no_price),
                "total_slippage_fiat": sum(abs(safe_float(r.get("Slippage ($)", 0))) for r in slip_f),
            }

    return findings


def analyze_asset_roll_forward(wb):
    """Analyze Asset Roll Forward report."""
    findings = {"report_type": "ASSET_ROLL_FORWARD", "report_name": "Asset Roll Forward"}

    headers, data = read_sheet_as_dicts(wb["raw_data"])
    if not data:
        findings["error"] = "No data in raw_data tab"
        return findings

    safety_col = next((h for h in headers if "Safety Check" in h), None)
    if safety_col:
        passing = [r for r in data if str(r.get(safety_col, "")).lower() in ("true", "1", "pass")]
        failing = [r for r in data if str(r.get(safety_col, "")).lower() in ("false", "0", "fail")]
        findings["safety_check"] = {
            "total_rows": len(data),
            "passing": len(passing),
            "failing": len(failing),
            "pass_rate": round(len(passing) / len(data) * 100, 1) if data else 0,
        }

    # Top assets by inflow/outflow
    inflow_col = next((h for h in headers if "Inflow" in h and "(T)" in h), None)
    outflow_col = next((h for h in headers if "Outflow" in h and "(T)" in h), None)

    if inflow_col:
        top_inflows = sorted(data, key=lambda r: abs(safe_float(r.get(inflow_col, 0))), reverse=True)[:5]
        findings["top_inflows"] = [
            {"asset": r.get("Asset Name", "?"), "wallet": r.get("Wallet Name", "?"),
             "amount": safe_float(r.get(inflow_col, 0))} for r in top_inflows
        ]

    if outflow_col:
        top_outflows = sorted(data, key=lambda r: abs(safe_float(r.get(outflow_col, 0))), reverse=True)[:5]
        findings["top_outflows"] = [
            {"asset": r.get("Asset Name", "?"), "wallet": r.get("Wallet Name", "?"),
             "amount": safe_float(r.get(outflow_col, 0))} for r in top_outflows
        ]

    return findings


def analyze_realized_gains(wb):
    """Analyze Realized Gains & Losses report."""
    findings = {"report_type": "EXTENDED_RAW_TRANSACTIONS", "report_name": "Realized Gains & Losses"}

    # Summary Per Year
    if "Summary Per Year" in wb.sheetnames:
        headers, data = read_sheet_as_dicts(wb["Summary Per Year"])
        findings["per_year"] = [
            {k: v for k, v in r.items()} for r in data
        ] if data else []

    # Summary Per Asset
    if "Summary Per Asset" in wb.sheetnames:
        headers, data = read_sheet_as_dicts(wb["Summary Per Asset"])
        if data:
            gain_col = next((h for h in headers if "Realized Gain" in h), None)
            if gain_col:
                sorted_by_gain = sorted(data, key=lambda r: safe_float(r.get(gain_col, 0)), reverse=True)
                findings["top_gains"] = [
                    {"asset": r.get("Asset Name", r.get("Asset Symbol", "?")),
                     "gain": safe_float(r.get(gain_col, 0))} for r in sorted_by_gain[:5]
                ]
                sorted_by_loss = sorted(data, key=lambda r: safe_float(r.get(gain_col, 0)))
                findings["top_losses"] = [
                    {"asset": r.get("Asset Name", r.get("Asset Symbol", "?")),
                     "gain": safe_float(r.get(gain_col, 0))} for r in sorted_by_loss[:5] if safe_float(r.get(gain_col, 0)) < 0
                ]

    # raw_data stats
    if "raw_data" in wb.sheetnames:
        headers, data = read_sheet_as_dicts(wb["raw_data"])
        if data:
            findings["raw_data"] = {
                "total_transactions": len(data),
                "unique_assets": len(set(r.get("Asset Name", r.get("Asset Symbol", "")) for r in data)),
            }
            # Date range
            date_col = next((h for h in headers if "Timestamp" in h or "Date" in h), None)
            if date_col:
                dates = [r.get(date_col) for r in data if r.get(date_col)]
                if dates:
                    findings["raw_data"]["date_range"] = {"min": str(min(dates)), "max": str(max(dates))}

            # Total realized gains
            gain_col = next((h for h in headers if "Realized Gain" in h), None)
            if gain_col:
                total = sum(safe_float(r.get(gain_col, 0)) for r in data)
                findings["raw_data"]["total_realized_gains"] = round(total, 2)

    return findings


def analyze_balances(wb, report_type):
    """Analyze Asset Balances / V2 / Historical Balance Format / Balance Trends."""
    name_map = {
        "RAW_BALANCES": "Asset Balances",
        "RAW_BALANCES_V2": "Asset Balances V2",
        "HISTORICAL_BALANCE": "Historical Balance Format",
        "BALANCE_TRENDS": "Balance Trends",
    }
    findings = {"report_type": report_type, "report_name": name_map.get(report_type, report_type)}

    if "raw_data" not in wb.sheetnames:
        findings["error"] = "No raw_data tab found"
        return findings

    headers, data = read_sheet_as_dicts(wb["raw_data"])
    if not data:
        findings["error"] = "No data in raw_data"
        return findings

    fiat_col = next((h for h in headers if "Fiat Value" in h and "Previous" not in h), None)
    amount_col = next((h for h in headers if h in ("Amount (T)", "Amount", "Token Amount")), None)

    findings["overview"] = {
        "total_rows": len(data),
        "unique_assets": len(set(r.get("Asset Name", r.get("Asset Symbol", "")) for r in data)),
        "unique_wallets": len(set(r.get("Wallet Name", r.get("Wallet Address", "")) for r in data)),
    }

    if fiat_col:
        total_fiat = sum(safe_float(r.get(fiat_col, 0)) for r in data)
        findings["overview"]["total_fiat_value"] = round(total_fiat, 2)

        top_holdings = sorted(data, key=lambda r: safe_float(r.get(fiat_col, 0)), reverse=True)[:10]
        findings["top_holdings"] = [
            {"asset": r.get("Asset Name", r.get("Asset Symbol", "?")),
             "wallet": r.get("Wallet Name", "?"),
             "fiat_value": safe_float(r.get(fiat_col, 0))} for r in top_holdings
        ]

    # Negative balances
    if amount_col:
        negatives = [r for r in data if safe_float(r.get(amount_col, 0)) < 0]
        if negatives:
            findings["negative_balances"] = len(negatives)

    # Missing prices
    no_price = [r for r in data if fiat_col and safe_float(r.get(fiat_col, 0)) == 0
                and amount_col and safe_float(r.get(amount_col, 0)) > 0]
    if no_price:
        findings["missing_prices"] = len(no_price)

    # Balance Trends specific
    if report_type == "BALANCE_TRENDS":
        change_col = next((h for h in headers if "Fiat Value Change" in h or "Fiat Change" in h), None)
        if not change_col:
            change_col = next((h for h in headers if "Change" in h and "$" in h), None)
        if change_col:
            sorted_up = sorted(data, key=lambda r: safe_float(r.get(change_col, 0)), reverse=True)[:5]
            sorted_down = sorted(data, key=lambda r: safe_float(r.get(change_col, 0)))[:5]
            findings["biggest_gainers"] = [
                {"asset": r.get("Asset Name", "?"), "change": safe_float(r.get(change_col, 0))} for r in sorted_up
            ]
            findings["biggest_losers"] = [
                {"asset": r.get("Asset Name", "?"), "change": safe_float(r.get(change_col, 0))} for r in sorted_down
            ]

    return findings


def analyze_cost_basis_stack(wb):
    """Analyze Cost Basis Stack Per Wallet."""
    findings = {"report_type": "COST_BASIS_STACK_PER_ACCOUNT", "report_name": "Cost Basis Stack Per Wallet"}

    headers, data = read_sheet_as_dicts(wb["raw_data"])
    if not data:
        findings["error"] = "No data"
        return findings

    findings["overview"] = {"total_lots": len(data)}

    # Unrealized gains
    ug_col = next((h for h in headers if "Unrealized Gain" in h), None)
    if ug_col:
        total_ug = sum(safe_float(r.get(ug_col, 0)) for r in data)
        findings["overview"]["total_unrealized_gains"] = round(total_ug, 2)
        losses = sorted([r for r in data if safe_float(r.get(ug_col, 0)) < 0],
                       key=lambda r: safe_float(r.get(ug_col, 0)))[:5]
        findings["tax_loss_candidates"] = [
            {"asset": r.get("Asset Name", "?"), "wallet": r.get("Wallet Name", "?"),
             "unrealized_loss": safe_float(r.get(ug_col, 0))} for r in losses
        ]

    # Impairment
    imp_col = next((h for h in headers if "Impairment" in h), None)
    if imp_col:
        total_imp = sum(abs(safe_float(r.get(imp_col, 0))) for r in data)
        findings["overview"]["total_impairment"] = round(total_imp, 2)

    return findings


def analyze_erp_pre_sync(wb):
    """Analyze ERP Pre-Sync."""
    findings = {"report_type": "PRE_SYNC_JOURNAL", "report_name": "ERP Pre-Sync"}

    headers, data = read_sheet_as_dicts(wb["raw_data"])
    if not data:
        findings["error"] = "No data"
        return findings

    status_col = next((h for h in headers if "Configuration Status" in h or "Config" in h), None)
    if status_col:
        from collections import Counter
        status_counts = Counter(str(r.get(status_col, "Unknown")) for r in data)
        findings["configuration_status"] = dict(status_counts)

    findings["total_entries"] = len(data)
    return findings


def analyze_erp_post_sync(wb):
    """Analyze ERP Post-Sync."""
    findings = {"report_type": "POST_SYNC_JOURNAL", "report_name": "ERP Post-Sync"}

    headers, data = read_sheet_as_dicts(wb["raw_data"])
    if not data:
        findings["error"] = "No data"
        return findings

    status_col = next((h for h in headers if "Sync Status" in h or "Status" in h), None)
    if status_col:
        from collections import Counter
        status_counts = Counter(str(r.get(status_col, "Unknown")) for r in data)
        findings["sync_status"] = dict(status_counts)
        failed = [r for r in data if "fail" in str(r.get(status_col, "")).lower()]
        if failed:
            findings["failed_syncs"] = len(failed)

    findings["total_entries"] = len(data)
    return findings


def analyze_generic(wb, report_type, report_name):
    """Generic analysis for reports without specialized logic."""
    findings = {"report_type": report_type, "report_name": report_name}

    if "raw_data" in wb.sheetnames:
        headers, data = read_sheet_as_dicts(wb["raw_data"])
        findings["total_rows"] = len(data)
        findings["columns"] = len(headers)
        findings["tabs"] = wb.sheetnames
    else:
        findings["tabs"] = wb.sheetnames
        # Try first sheet
        ws = wb[wb.sheetnames[0]]
        findings["total_rows"] = ws.max_row - 1 if ws.max_row else 0

    return findings


def analyze(xlsx_path):
    """Main analysis entry point."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    report_type = identify_report_type(wb)

    analyzers = {
        "RECONCILIATION_LEDGER": lambda: analyze_reconciliation_ledger(wb),
        "ASSET_ROLL_FORWARD": lambda: analyze_asset_roll_forward(wb),
        "EXTENDED_RAW_TRANSACTIONS": lambda: analyze_realized_gains(wb),
        "RAW_BALANCES": lambda: analyze_balances(wb, "RAW_BALANCES"),
        "RAW_BALANCES_V2": lambda: analyze_balances(wb, "RAW_BALANCES_V2"),
        "HISTORICAL_BALANCE": lambda: analyze_balances(wb, "HISTORICAL_BALANCE"),
        "BALANCE_TRENDS": lambda: analyze_balances(wb, "BALANCE_TRENDS"),
        "COST_BASIS_STACK_PER_ACCOUNT": lambda: analyze_cost_basis_stack(wb),
        "PRE_SYNC_JOURNAL": lambda: analyze_erp_pre_sync(wb),
        "POST_SYNC_JOURNAL": lambda: analyze_erp_post_sync(wb),
    }

    if report_type in analyzers:
        findings = analyzers[report_type]()
    else:
        name_map = {
            "BASIC_RAW_TRANSACTIONS": "Transaction Ledger",
            "COST_BREAKDOWN_RAW_TRANSACTIONS": "Cost Breakdown",
            "ROLLUP_BREAKDOWN": "Rollup Breakdown",
            "ARCHIVED_BALANCES": "Asset Balances - Archives",
            "HISTORICAL_BALANCE": "Historical Balance Format",
            "COST_BASIS_INVENTORY": "Cost Basis Inventory",
            "COST_BASIS_ROLL_FORWARD": "Cost Basis Roll Forward",
            "DAILY_ASSET_PRICING": "Asset Fiat Values",
            "INTERNAL_ACCOUNTS": "Organization Wallets",
            "REEVALUATION": "Revaluation",
        }
        findings = analyze_generic(wb, report_type, name_map.get(report_type, "Unknown"))

    wb.close()
    return findings


def main():
    parser = argparse.ArgumentParser(description="Analyze a TRES Finance report XLSX")
    parser.add_argument("xlsx_path", help="Path to the XLSX file")
    parser.add_argument("--output", "-o", help="Output JSON path", default=None)
    args = parser.parse_args()

    findings = analyze(args.xlsx_path)

    output = json.dumps(findings, indent=2, default=str)

    if args.output:
        Path(args.output).parent.mkdir(parents=True, exist_ok=True)
        with open(args.output, "w") as f:
            f.write(output)
        print(f"Analysis saved to {args.output}")
    else:
        print(output)


if __name__ == "__main__":
    main()
